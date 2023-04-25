"""
Module: contains class ExcelParser.
"""
import os
from typing import Union

import openpyxl
import pandas

from .regex_parser import RegexParser


class ExcelParser:
    """
    Reads existing Excel spreadsheet, parses desired targets & adds columns.
    """

    def __init__(
        self, excel_filename: str, sheet_name: Union[str, None] = None
    ) -> None:
        """Reads in one sheet of the Excel file.

        Parameters
        ----------
        excel_filename : str        Name of existing Excel file
        sheet_name : Optional str   If not specified, reads first sheet.
        """
        if not isinstance(excel_filename, str):
            raise TypeError("Argument 'excel_filename' is not the expected str.")

        if not os.path.isfile(excel_filename):
            raise FileNotFoundError(f"Unable to find file '{excel_filename}'.")

        self.__excel_filename = excel_filename

        if not isinstance(sheet_name, str):
            #   Look up active sheet name.
            wb = openpyxl.load_workbook(self.__excel_filename)
            sheet_name = wb.active.title
            wb.close()

        self.__sheet_name = sheet_name
        df: pandas.DataFrame = pandas.read_excel(
            self.__excel_filename, sheet_name=sheet_name
        )

        if not isinstance(df, pandas.DataFrame):  # pragma: no cover
            raise RuntimeError(f"Unable to read file '{self.__excel_filename}'.")

        #   Farm out the Regular Expression work to the RegexParser class,
        #   so it can be reused with other file types.
        self.__regex_parser: RegexParser = RegexParser(df)

    def clean_column(self, column_name: str, pattern: str, replace: str) -> None:
        """Use a regex to fix strings.

        Parameters
        ----------
        column_name : str
        pattern : str
        replace : str
        """
        self.__regex_parser.clean_column(
            column_name=column_name, pattern=pattern, replace=replace
        )

    def data(self) -> pandas.DataFrame:
        """Allows read access to self.__df.

        Returns
        -------
        df : pandas.DataFrame
        """
        return self.__regex_parser.data()

    def extract(self, column_name: str, pattern: Union[str, list]) -> list:
        """Use a regex to extract data from a given column into a list.

        Parameters
        ----------
        column_name : str
        pattern : str

        Returns
        -------
        extracted_data : list of str
        """
        return self.__regex_parser.extract(column_name=column_name, pattern=pattern)

    def extract_into_new_column(
        self, column_name: str, pattern: Union[str, list], new_column: str
    ) -> None:
        """Use a regex to extract data from a given column into a new column.

        Parameters
        ----------
        column_name : str
        pattern : str or list of str
        new_column : str
        """
        self.__regex_parser.extract_into_new_column(
            column_name=column_name, pattern=pattern, new_column=new_column
        )

    def restore_original_column(self, column_name: str) -> None:
        """Restores the original (uncleaned) column in preparation for writing out results.
        In (optional) cleaning, we may have changed the source column in the dataframe.
        But when writing out the results, we want to show the original column.

        Parameters
        ----------
        column_name : str
        """
        self.__regex_parser.restore_original_column(column_name=column_name)

    def write_to_excel(self, new_file_name: Union[str, None] = None) -> str:
        """Write out the dataframe we've been building.

        Parameters
        ----------
        new_file_name : Optional str

        Returns
        -------
        new_file_name : str
        """
        if not new_file_name:
            name, extension = os.path.splitext(os.path.basename(self.__excel_filename))
            new_file_name = os.path.join(
                os.path.dirname(self.__excel_filename), name + "_revised" + extension
            )

        # https: // stackoverflow.com / a / 72446796 / 18749636
        wb_obj = openpyxl.Workbook()
        wb_obj.save(new_file_name)
        sheet = wb_obj.active
        sheet.title = self.__sheet_name
        start_col = 1
        start_row = 1
        col_idx = start_col
        df = self.__regex_parser.data()

        # insert values
        for label, content in df.items():
            sheet.cell(row=start_row, column=col_idx, value=label)

            for row_idx, value_ in enumerate(content):
                sheet.cell(row=start_row + row_idx + 1, column=col_idx, value=value_)

            col_idx += 1

        wb_obj.save(new_file_name)
        return new_file_name
