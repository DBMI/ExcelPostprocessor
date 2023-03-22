"""
Moodule: contains class ExcelParser.
"""
import os
from typing import Union

import openpyxl
import pandas


class ExcelParser:
    """
    Reads existing Excel spreadsheet, parses desired targets & adds columns.
    """

    def __init__(self, excel_filename: str, sheet_name: Union[str, None] = None):
        """Reads in one sheet of the Excel file.

        Parameters
        ----------
        excel_filename : str        Name of existing Excel file
        sheet_name : Optional str   If not specified, reads first sheet.
        """
        if not isinstance(excel_filename, str):
            raise TypeError("Argument 'excel_filename' is not the expected str.")

        if not os.path.isfile(excel_filename):
            raise FileNotFoundError(f"Unable to find file '{excel_filename}'.")

        self.__excel_filename = excel_filename

        if not isinstance(sheet_name, str):
            #   Look up active sheet name.
            wb = openpyxl.load_workbook(self.__excel_filename)
            sheet_name = wb.active.title
            wb.close()

        self.__sheet_name = sheet_name
        self.__df = pandas.read_excel(self.__excel_filename, sheet_name=sheet_name)

        if not isinstance(self.__df, pandas.DataFrame):  # pragma: no cover
            raise RuntimeError(f"Unable to read file '{self.__excel_filename}'.")

    def data(self) -> pandas.DataFrame:
        """Allows read access to self.__df.

        Returns
        -------
        df : pandas.DataFrame
        """
        return self.__df

    def extract(self, column_name: str, pattern: str) -> list:
        """Use a regex to extract data from a given column into a list.

        Parameters
        ----------
        column_name : str
        pattern : str

        Returns
        -------
        extracted_data : list of str
        """
        extracted_data = self.__extract_series(column_name=column_name, pattern=pattern)
        return extracted_data.tolist()

    def __extract_series(self, column_name: str, pattern: str) -> pandas.Series:
        """Extracts column to Series using regex.

        Parameters
        ----------
        column_name : str
        pattern : str

        Returns
        -------
        column : Series
        """
        if not isinstance(column_name, str):
            raise TypeError("Argument 'column_name' is not the expected str.")

        if not isinstance(pattern, str):
            raise TypeError("Argument 'pattern' is not the expected str.")

        if column_name not in self.__df:
            raise AttributeError(f"Unable to find column '{column_name}' in DataFrame.")

        column: pandas.Series = self.__df[column_name].str.extract(pattern).squeeze()
        return column

    def extract_into_new_column(
        self, column_name: str, pattern: str, new_column: str
    ) -> None:
        """Use a regex to extract data from a given column into a new column.

        Parameters
        ----------
        column_name : str
        pattern : str
        new_column : str
        """
        extracted_data = self.__extract_series(column_name=column_name, pattern=pattern)
        self.__df[new_column] = extracted_data

    def write_to_excel(self, new_file_name: Union[str, None] = None) -> str:
        """Write out the dataframe we've been building.

        Parameters
        ----------
        new_file_name : Optional str

        Returns
        -------
        new_file_name : str
        """
        if not new_file_name:
            name, extension = os.path.splitext(os.path.basename(self.__excel_filename))
            new_file_name = os.path.join(
                os.path.dirname(self.__excel_filename), name + "_revised" + extension
            )

        # https: // stackoverflow.com / a / 72446796 / 18749636
        wb_obj = openpyxl.Workbook()
        wb_obj.save(new_file_name)
        sheet = wb_obj.active
        sheet.title = self.__sheet_name
        start_col = 1
        start_row = 1
        col_idx = start_col

        # insert values
        for label, content in self.__df.items():
            sheet.cell(row=start_row, column=col_idx, value=label)

            for row_idx, value_ in enumerate(content):
                sheet.cell(row=start_row + row_idx + 1, column=col_idx, value=value_)

            col_idx += 1

        wb_obj.save(new_file_name)
        return new_file_name
